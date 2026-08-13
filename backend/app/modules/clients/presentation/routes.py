from fastapi import APIRouter, Query, Response
from pydantic import BaseModel

from app.core.deps import CurrentUser, DbSession
from app.modules.clients.application.service import ClientService
from app.modules.clients.infrastructure.repositories import (
    ClientContactRepository,
    ClientLocationRepository,
    ClientRepository,
    PatientRepository,
    ProjectRepository,
)
from app.modules.clients.presentation.schemas import (
    ClientCreateRequest,
    ClientListResponse,
    ClientUpdateRequest,
    ContactCreateRequest,
    LocationCreateRequest,
    PatientCreateRequest,
)

router = APIRouter(prefix="/clients", tags=["Clients"])


def get_service(db: DbSession) -> ClientService:
    return ClientService(
        client_repo=ClientRepository(db), patient_repo=PatientRepository(db),
        project_repo=ProjectRepository(db), contact_repo=ClientContactRepository(db),
        location_repo=ClientLocationRepository(db),
    )


import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

@router.get("/template")
async def download_clients_template():
    wb = openpyxl.Workbook()

    # ── Sheet 1: Data Entry Sheet ─────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Carga Clientes y Sedes"

    headers_display = [
        "NIT / Cédula (*)",
        "Nombre / Razón Social (*)",
        "Nombre Comercial / Sede",
        "Tipo de Cliente (*)",
        "Correo Electrónico",
        "Teléfono Fijo",
        "Celular / Móvil",
        "Dirección",
        "Departamento (*)",
        "Ciudad / Municipio (*)",
        "Latitud GPS",
        "Longitud GPS",
        "Radio Geocerca (Metros)",
        "Observaciones / Notas"
    ]

    header_req_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")  # Dark Slate Blue
    header_opt_fill = PatternFill(start_color="475569", end_color="475569", fill_type="solid")  # Muted Slate
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    border_thin = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws1.append(headers_display)
    for col_idx, cell in enumerate(ws1[1], start=1):
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border_thin
        # Highlight required columns (*) in dark slate, optional in slate
        if "(*)" in headers_display[col_idx - 1]:
            cell.fill = header_req_fill
        else:
            cell.fill = header_opt_fill

    sample_rows = [
        ["901540816-8", "Home Care del Quindío IPS", "Sede Principal Armenia", "IPS", "atencion@homecarequindio.com", "6067412345", "3234790310", "Carrera 14 # 23-45", "Quindío", "Armenia", 4.5389, -75.6725, 100, "Sede principal de atención médica IPS"],
        ["900123456-1", "Edificio Torre Central", "Torre Operativa Bogotá", "Empresa", "contacto@torrecentral.com", "6015551234", "3109876543", "Calle 100 # 19-61", "Bogotá D.C.", "Bogotá D.C.", 4.6835, -74.0532, 150, "Oficinas corporativas del cliente"],
        ["890123789-0", "Hospital Universitario San Juan", "Sede Urgencias Norte", "Hospital", "contacto@hospitalsanjuan.gov.co", "6044445566", "3158889900", "Avenida 4 Norte # 12-34", "Valle del Cauca", "Cali", 3.4516, -76.5320, 120, "Centro asistencial hospitalario"],
        ["901999888-5", "Constructora del Café S.A.S.", "Proyecto Altamira Residencial", "Proyecto", "obras@constructoradelcafe.com", "6067332211", "3124443322", "Km 3 Vía a Montenegro", "Quindío", "Montenegro", 4.5620, -75.7480, 200, "Obra de construcción en desarrollo"],
    ]

    for row_data in sample_rows:
        ws1.append(row_data)
        for cell in ws1[ws1.max_row]:
            cell.border = border_thin
            cell.alignment = Alignment(vertical="center")

    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws1.column_dimensions[col_letter].width = max(max_len + 4, 16)
    ws1.row_dimensions[1].height = 30

    # ── Sheet 2: Instructions & Guide Sheet ──────────────────────────────────
    ws2 = wb.create_sheet(title="Guía e Instrucciones")

    ws2.append(["GUÍA DE CARGA MASIVA DE CLIENTES Y SEDES - DEACONTROL ENTERPRISE"])
    ws2.append([])

    guide_headers = ["Nombre del Campo", "Obligatorio", "Valores Permitidos / Formato", "Descripción y Ejemplo"]
    ws2.append(guide_headers)

    g_header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    for cell in ws2[3]:
        cell.fill = g_header_fill
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.border = border_thin

    guide_rows = [
        ["NIT / Cédula (*)", "SÍ", "Texto (ej: 900123456-1)", "Número de documento de identificación del cliente o sede."],
        ["Nombre / Razón Social (*)", "SÍ", "Texto libre", "Razón social oficial o nombre del cliente. Ej: Home Care del Quindío IPS"],
        ["Nombre Comercial / Sede", "NO", "Texto libre", "Nombre específico de la sede o nombre comercial. Ej: Sede Principal Armenia"],
        ["Tipo de Cliente (*)", "SÍ", "Empresa, Persona Natural, IPS, Hospital, Clínica, Proyecto", "Categoría del cliente para clasificación en el sistema."],
        ["Correo Electrónico", "NO", "Email válido", "Correo de contacto corporativo. Ej: contacto@empresa.com"],
        ["Teléfono Fijo", "NO", "Números (ej: 6067412345)", "Teléfono de contacto con indicativo de ciudad."],
        ["Celular / Móvil", "NO", "10 dígitos (ej: 3109876543)", "Número celular de contacto directo."],
        ["Dirección", "NO", "Texto libre", "Dirección física de la sede o cliente. Ej: Carrera 14 # 23-45"],
        ["Departamento (*)", "SÍ", "Departamento oficial CO (ej: Quindío)", "Departamento geográfico de ubicación."],
        ["Ciudad / Municipio (*)", "SÍ", "Municipio oficial CO (ej: Armenia)", "Municipio geográfico de ubicación."],
        ["Latitud GPS", "NO", "Decimal (ej: 4.5389)", "Coordenada de latitud GPS para control de geocercas en marcado asistido."],
        ["Longitud GPS", "NO", "Decimal (ej: -75.6725)", "Coordenada de longitud GPS para marcación en app móvil."],
        ["Radio Geocerca (Metros)", "NO", "Número entero (ej: 100)", "Distancia máxima en metros permitida para marcación. Por defecto 100m."],
        ["Observaciones / Notas", "NO", "Texto libre", "Notas internas o descripción operativa de la sede."],
    ]

    for r_data in guide_rows:
        ws2.append(r_data)
        row_cells = ws2[ws2.max_row]
        for idx, cell in enumerate(row_cells):
            cell.border = border_thin
            if idx == 1:
                cell.alignment = Alignment(horizontal="center")
                if r_data[1] == "SÍ":
                    cell.font = Font(bold=True, color="DC2626")  # Red for required
                else:
                    cell.font = Font(color="64748B")
            elif idx == 0:
                cell.font = Font(bold=True)

    for col in ws2.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws2.column_dimensions[col_letter].width = max(max_len + 3, 20)

    ws2.row_dimensions[1].height = 24
    ws2[1][0].font = Font(size=14, bold=True, color="1E3A8A")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_clientes_sedes_deacontrol.xlsx"},
    )


class ClientImportRequest(BaseModel):
    company_id: str | None = None
    clients: list[dict]


@router.post("/import")
async def import_clients_bulk(body: ClientImportRequest, current_user: CurrentUser, db: DbSession):
    service = get_service(db)
    return await service.bulk_import_clients(body.company_id or current_user.company_id, body.clients)


@router.post("", status_code=201)
async def create_client(body: ClientCreateRequest, current_user: CurrentUser, db: DbSession) -> dict:
    data = body.model_dump()
    if not data.get("company_id"):
        data["company_id"] = getattr(current_user, "company_id", None) or "dla-company-main"
    res = await get_service(db).create_client(**data)
    await db.commit()
    return res


@router.get("", response_model=ClientListResponse)
async def list_clients(
    current_user: CurrentUser, db: DbSession,
    company_id: str | None = Query(None), client_type: str | None = Query(None),
    status: str | None = Query(None), search: str | None = Query(None),
    page: int = Query(1, ge=1), page_size: int = Query(25, ge=1, le=1000),
) -> ClientListResponse:
    result = await get_service(db).list_clients(
        company_id=company_id, client_type=client_type, status=status,
        search=search, page=page, page_size=page_size,
    )
    return ClientListResponse(**result.__dict__)


@router.get("/{client_id}")
async def get_client(client_id: str, current_user: CurrentUser, db: DbSession) -> dict:
    return await get_service(db).get_client(client_id)


@router.put("/{client_id}")
async def update_client(client_id: str, body: ClientUpdateRequest, current_user: CurrentUser, db: DbSession) -> dict:
    return await get_service(db).update_client(client_id, **body.model_dump(exclude_unset=True))


@router.delete("/{client_id}")
async def delete_client(client_id: str, current_user: CurrentUser, db: DbSession) -> dict:
    res = await get_service(db).delete_client(client_id, db=db)
    await db.commit()
    return res


@router.patch("/{client_id}/status")
async def update_client_status(client_id: str, body: dict, current_user: CurrentUser, db: DbSession) -> dict:
    new_status = body.get("status", "inactive")
    return await get_service(db).update_status(client_id, new_status)


@router.get("/{client_id}/contacts")
async def list_contacts(client_id: str, current_user: CurrentUser, db: DbSession) -> list[dict]:
    return await get_service(db).list_contacts(client_id)


@router.post("/{client_id}/contacts", status_code=201)
async def add_contact(client_id: str, body: ContactCreateRequest, current_user: CurrentUser, db: DbSession) -> dict:
    return await get_service(db).add_contact(client_id, **body.model_dump())


@router.delete("/{client_id}/contacts/{contact_id}")
async def delete_contact(client_id: str, contact_id: str, current_user: CurrentUser, db: DbSession) -> dict:
    return await get_service(db).delete_contact(client_id, contact_id)


@router.get("/{client_id}/locations")
async def list_locations(client_id: str, current_user: CurrentUser, db: DbSession) -> list[dict]:
    return await get_service(db).list_locations(client_id)


@router.post("/{client_id}/locations", status_code=201)
async def add_location(client_id: str, body: LocationCreateRequest, current_user: CurrentUser, db: DbSession) -> dict:
    return await get_service(db).add_location(client_id, **body.model_dump())


@router.delete("/{client_id}/locations/{location_id}")
async def delete_location(client_id: str, location_id: str, current_user: CurrentUser, db: DbSession) -> dict:
    return await get_service(db).delete_location(client_id, location_id)


@router.get("/{client_id}/patients")
async def list_patients(
    client_id: str, current_user: CurrentUser, db: DbSession,
    search: str | None = Query(None), page: int = Query(1, ge=1), page_size: int = Query(25),
) -> ClientListResponse:
    result = await get_service(db).list_patients(client_id, search=search, page=page, page_size=page_size)
    return ClientListResponse(**result.__dict__)


@router.post("/{client_id}/patients", status_code=201)
async def create_patient(client_id: str, body: PatientCreateRequest, current_user: CurrentUser, db: DbSession) -> dict:
    return await get_service(db).create_patient(client_id, **body.model_dump())


@router.delete("/{client_id}/patients/{patient_id}")
async def delete_patient(client_id: str, patient_id: str, current_user: CurrentUser, db: DbSession) -> dict:
    return await get_service(db).delete_patient(client_id, patient_id, db=db)


@router.post("/{client_id}/projects", status_code=201)
async def create_project(client_id: str, code: str, name: str, current_user: CurrentUser, db: DbSession) -> dict:
    return await get_service(db).create_project(client_id, code=code, name=name)


@router.get("/{client_id}/personas")
async def list_personas(
    client_id: str, current_user: CurrentUser, db: DbSession,
    search: str | None = Query(None), page: int = Query(1, ge=1), page_size: int = Query(25),
):
    from sqlalchemy import select, or_, func
    from app.shared.database.models_clients import Persona
    stmt = select(Persona).where(Persona.client_id == client_id, Persona.is_deleted == False)
    if search:
        stmt = stmt.where(Persona.first_name.ilike(f"%{search}%") | Persona.last_name.ilike(f"%{search}%"))
    result = await db.execute(stmt.offset((page - 1) * page_size).limit(page_size))
    personas = result.scalars().all()
    count_stmt = select(func.count(Persona.id)).where(Persona.client_id == client_id, Persona.is_deleted == False)
    total = (await db.execute(count_stmt)).scalar() or 0
    return {"items": [
        {
            "id": p.id, "first_name": p.first_name, "last_name": p.last_name,
            "document_type": p.document_type, "document_number": p.document_number,
            "email": p.email, "phone": p.phone, "status": p.status,
        }
        for p in personas
    ], "total": total, "page": page, "page_size": page_size}
