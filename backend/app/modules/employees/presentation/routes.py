from fastapi import APIRouter, Query, Response
from pydantic import BaseModel

from app.core.deps import CurrentUser, DbSession
from app.modules.employees.application.service import EmployeeService
from app.modules.employees.infrastructure.repositories import (
    EmployeeDocumentRepository,
    EmployeeRepository,
)
from app.modules.employees.presentation.schemas import (
    EmployeeCreateRequest,
    EmployeeListResponse,
    EmployeeUpdateRequest,
)

router = APIRouter(prefix="/employees", tags=["Employees"])


def get_service(db: DbSession) -> EmployeeService:
    return EmployeeService(
        employee_repo=EmployeeRepository(db),
        document_repo=EmployeeDocumentRepository(db),
    )


@router.post("", status_code=201)
async def create_employee(body: EmployeeCreateRequest, current_user: CurrentUser, db: DbSession) -> dict:
    service = get_service(db)
    return await service.create_employee(**body.model_dump())


@router.get("", response_model=EmployeeListResponse)
async def list_employees(
    current_user: CurrentUser,
    db: DbSession,
    company_id: str | None = Query(None),
    department_id: str | None = Query(None),
    status: str | None = Query(None),
    search: str | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(100, ge=1, le=1000),
) -> EmployeeListResponse:
    service = get_service(db)
    result = await service.list_employees(
        company_id=company_id, department_id=department_id,
        status=status, search=search, page=page, page_size=page_size,
    )
    return EmployeeListResponse(**result.__dict__)


import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

@router.get("/template")
async def download_employees_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla Empleados"

    # Header columns
    headers = [
        "code", "document_type", "document_number", "first_name", "last_name",
        "email", "phone", "mobile", "job_position", "department", "eps", "arl", "afp", "status"
    ]
    
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid") # Dark Slate / Navy
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    border_thin = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border_thin

    # Sample rows for guidance
    sample_rows = [
        ["EMP-1001", "CC", "1020304050", "Carlos", "Pérez", "carlos.perez@empresa.com", "3001234567", "3001234567", "Operador de Campo", "Operaciones", "EPS Sura", "Positiva ARL", "Porvenir", "active"],
        ["EMP-1002", "CC", "1030405060", "María", "Gómez", "maria.gomez@empresa.com", "3109876543", "3109876543", "Supervisora de Sede", "Operaciones", "Sanitas", "SURA ARL", "Protección", "active"],
    ]

    for row_data in sample_rows:
        ws.append(row_data)
        for cell in ws[ws.max_row]:
            cell.border = border_thin
            cell.alignment = Alignment(vertical="center")

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

    ws.row_dimensions[1].height = 28

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_empleados_deacontrol.xlsx"},
    )


class EmployeeImportRequest(BaseModel):
    company_id: str | None = None
    employees: list[dict]


@router.post("/import")
async def import_employees_bulk(body: EmployeeImportRequest, current_user: CurrentUser, db: DbSession):
    service = get_service(db)
    return await service.bulk_import_employees(body.company_id or current_user.company_id, body.employees)


@router.get("/stats/summary")
async def get_employee_stats(company_id: str, current_user: CurrentUser, db: DbSession) -> dict:
    service = get_service(db)
    return await service.get_company_stats(company_id)


@router.get("/{employee_id}")
async def get_employee(employee_id: str, current_user: CurrentUser, db: DbSession) -> dict:
    service = get_service(db)
    employee = await service.get_employee(employee_id)
    return {
        "id": employee.id,
        "code": employee.code,
        "document_type": employee.document_type,
        "document_number": employee.document_number,
        "first_name": employee.first_name,
        "last_name": employee.last_name,
        "middle_name": employee.middle_name,
        "second_last_name": employee.second_last_name,
        "email": employee.email,
        "phone": employee.phone,
        "mobile": employee.mobile,
        "address": employee.address,
        "city": employee.city,
        "status": employee.status,
        "photo_url": employee.photo_url,
        "company_id": employee.company_id,
        "branch_id": employee.branch_id,
        "department_id": employee.department_id,
        "job_position_id": employee.job_position_id,
        "cost_center_id": employee.cost_center_id,
        "work_team_id": employee.work_team_id,
        "hire_date": str(employee.hire_date) if employee.hire_date else None,
        "birth_date": str(employee.birth_date) if employee.birth_date else None,
        "gender": employee.gender,
        "blood_type": employee.blood_type,
        "marital_status": employee.marital_status,
        "eps": employee.eps,
        "arl": employee.arl,
        "afp": employee.afp,
        "caja_compensacion": employee.caja_compensacion,
        "emergency_contact_name": employee.emergency_contact_name,
        "emergency_contact_phone": employee.emergency_contact_phone,
        "emergency_contact_relation": employee.emergency_contact_relation,
        "bank_name": employee.bank_name,
        "bank_account_type": employee.bank_account_type,
        "bank_account_number": employee.bank_account_number,
        "platform_access": employee.platform_access,
        "account_status": employee.account_status,
        "has_access": bool(employee.username),
        "username": employee.username,
        "role_id": employee.role_id,
    }


@router.put("/{employee_id}")
async def update_employee(employee_id: str, body: EmployeeUpdateRequest, current_user: CurrentUser, db: DbSession) -> dict:
    service = get_service(db)
    return await service.update_employee(employee_id, **body.model_dump(exclude_unset=True))


@router.delete("/{employee_id}")
async def delete_employee(employee_id: str, current_user: CurrentUser, db: DbSession) -> dict:
    service = get_service(db)
    res = await service.delete_employee(employee_id, db=db)
    await db.commit()
    return res


@router.post("/{employee_id}/access", status_code=201)
async def create_employee_access(employee_id: str, body: dict, current_user: CurrentUser, db: DbSession) -> dict:
    service = get_service(db)
    return await service.create_access(
        employee_id,
        username=body.get("username", ""),
        password=body.get("password", ""),
        role_id=body.get("role_id"),
        platform_access=body.get("platform_access", "both"),
    )


@router.put("/{employee_id}/access")
async def update_employee_access(employee_id: str, body: dict, current_user: CurrentUser, db: DbSession) -> dict:
    service = get_service(db)
    return await service.update_access(employee_id, **body)


@router.post("/{employee_id}/access/reset-password")
async def reset_employee_password(employee_id: str, body: dict, current_user: CurrentUser, db: DbSession) -> dict:
    service = get_service(db)
    return await service.reset_password(employee_id, body.get("password", ""))


@router.get("/{employee_id}/documents")
async def list_employee_documents(employee_id: str, current_user: CurrentUser, db: DbSession) -> list[dict]:
    service = get_service(db)
    return await service.get_employee_documents(employee_id)


@router.post("/{employee_id}/documents", status_code=201)
async def add_employee_document(employee_id: str, document_type: str, name: str, file_url: str, current_user: CurrentUser, db: DbSession) -> dict:
    service = get_service(db)
    return await service.add_employee_document(employee_id, document_type=document_type, name=name, file_url=file_url)
